"""
Lógica de processamento — separada da UI para facilitar testes.
"""
import io
import re
import csv
from datetime import datetime

import openpyxl


# ── leitura de arquivos ──────────────────────────────────────────────────────

def _padrow(row, n=9):
    """Garante que a linha tenha pelo menos n elementos."""
    return list(row) + [""] * max(0, n - len(row))


def ler_xlsx(file_obj):
    """Lê a primeira aba de um XLSX e devolve lista de linhas (lista de str)."""
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows():
        rows.append([str(c.value).strip() if c.value is not None else "" for c in row])
    wb.close()
    return rows


def ler_csv_diaria(file_obj):
    """Lê um CSV de diária (saída do sistema) tratando campos multi-linha."""
    text = file_obj.read().decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def ler_arquivo_diaria(file_obj, nome_arquivo):
    """Detecta formato pelo nome e devolve linhas."""
    ext = nome_arquivo.rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xls"):
        return ler_xlsx(file_obj)
    return ler_csv_diaria(file_obj)


# ── helpers ──────────────────────────────────────────────────────────────────

def extrair_cpf(nome):
    m = re.search(r"[-\s]+(\d{11,14})\s*$", nome.strip())
    return m.group(1).strip() if m else ""


def extrair_nome_credor(credor):
    """Do campo 'FULANO DE TAL - 12345678901' devolve só o nome (sem o CPF)."""
    s = str(credor).strip()
    # remove o sufixo " - CPF" (11 a 14 dígitos)
    s = re.sub(r"[-\s]+\d{11,14}\s*$", "", s)
    return s.strip()


def mes_ano(data_str):
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(data_str.strip(), fmt).strftime("%m/%Y")
        except ValueError:
            pass
    return data_str.strip()


def normalizar_valor(v):
    if isinstance(v, (int, float)):
        return f"{round(float(v), 2):.2f}"
    s = str(v).strip().replace("\n", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return f"{round(float(s), 2):.2f}"
    except ValueError:
        return s


# ── processamento de diárias ─────────────────────────────────────────────────

SITUACOES_IGNORAR = {"", "SITUAÇÃO", "SITUACAO", "SITUA?O", "NONE", "SITUAÃ§ÃƒO"}
SITUACOES_ANULADO = {"ANULADO", "REJEITADA"}


def processar_diarias(rows):
    """
    Extrai registros válidos de um relatório de diárias.
    Ignora linhas de subtotal, cabeçalho e situações ANULADO/REJEITADA.
    """
    registros = []
    for r in rows:
        r = _padrow(r, 9)
        situacao = r[2].strip().upper()

        if situacao in SITUACOES_IGNORAR:
            continue
        if situacao in SITUACOES_ANULADO:
            continue

        cpf = extrair_cpf(r[0])
        ob  = r[3].strip()
        data = r[5].strip()
        valor = normalizar_valor(r[8])

        # descarta linhas sem OB ou sem CPF/CNPJ identificável
        if not ob or not cpf:
            continue

        # descarta valores negativos (estornos)
        try:
            if float(valor) < 0:
                continue
        except ValueError:
            pass

        registros.append({
            "CPF_CNPJ": cpf,
            "NOME":     extrair_nome_credor(r[0]),
            "OB":       ob,
            "DATA":     data,
            "VALOR":    valor,
        })
    return registros


# ── base de servidores ────────────────────────────────────────────────────────

def carregar_servidores_xlsx(file_obj):
    rows = ler_xlsx(file_obj)
    return _indexar_servidores(rows)


def carregar_servidores_csv(file_obj):
    text = file_obj.read().decode("utf-8", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    return _indexar_servidores(rows)


def _indexar_servidores(rows):
    """Devolve dict CPF -> {MATRICULA, VINCULO, NOME}."""
    servidores = {}
    for r in rows[1:]:
        r = _padrow(r, 5)
        if not r[0].strip().isdigit():
            continue
        cpf = r[4].strip().zfill(11)
        servidores[cpf] = {
            "MATRICULA": r[0].strip(),
            "VINCULO":   r[1].strip(),
            "NOME":      r[3].strip(),
        }
    return servidores


def servidores_para_csv(servidores: dict) -> bytes:
    """Serializa o dict de servidores de volta para CSV (para salvar no disco)."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["MATRICULA", "VINCULO", "MATVINC", "SERVIDOR", "CPF"])
    for cpf, srv in servidores.items():
        w.writerow([srv["MATRICULA"], srv["VINCULO"], "", srv["NOME"], cpf])
    return buf.getvalue().encode("utf-8-sig")


# ── cruzamento ────────────────────────────────────────────────────────────────

def cruzar(registros: list, servidores: dict):
    """
    Cruza registros de diárias com cadastro de servidores.
    Deduplica por OB.
    Retorna (encontrados, cpfs_nao_encontrados).
    """
    encontrados_map = {}
    nao_encontrados = set()

    for reg in registros:
        cpf = reg["CPF_CNPJ"]
        # normaliza CPF pessoa física para 11 dígitos
        if cpf.isdigit() and len(cpf) <= 11:
            cpf = cpf.zfill(11)

        ob  = reg["OB"]
        srv = servidores.get(cpf)

        if srv:
            if ob not in encontrados_map:
                encontrados_map[ob] = {
                    "MATRICULA": srv["MATRICULA"],
                    "VINCULO":   srv["VINCULO"],
                    "NOME":      srv["NOME"],
                    "OB":        ob,
                    "MES_ANO":   mes_ano(reg["DATA"]),
                    "VALOR":     reg["VALOR"],
                }
        else:
            # só reporta CPFs de pessoa física (11 dígitos) como não encontrados
            if cpf.isdigit() and len(cpf) == 11:
                nao_encontrados.add(cpf)

    resultado = sorted(encontrados_map.values(), key=lambda x: x["NOME"])
    return resultado, nao_encontrados


# ── estornos pendentes ────────────────────────────────────────────────────────

def verificar_estornos(resultado: list) -> list:
    """
    Para cada OB com valor negativo, verifica se existe outra OB da mesma
    matrícula+vínculo com o valor positivo correspondente.
    Retorna lista de dicts com os estornos que NÃO foram repostos.
    """
    # índice: (matricula, vinculo) -> lista de valores positivos
    positivos: dict = {}
    for r in resultado:
        v = float(r["VALOR"])
        if v > 0:
            chave = (r["MATRICULA"], r["VINCULO"])
            positivos.setdefault(chave, []).append(round(v, 2))

    pendentes = []
    for r in resultado:
        v = float(r["VALOR"])
        if v < 0:
            chave = (r["MATRICULA"], r["VINCULO"])
            valor_positivo = round(abs(v), 2)
            lista = positivos.get(chave, [])
            if valor_positivo not in lista:
                pendentes.append({
                    "MATRICULA": r["MATRICULA"],
                    "VINCULO":   r["VINCULO"],
                    "NOME":      r["NOME"],
                    "OB_ESTORNO": r["OB"],
                    "MES_ANO":   r["MES_ANO"],
                    "VALOR_ESTORNADO": r["VALOR"],
                    "SITUACAO": "⚠️ SEM REPOSIÇÃO — VERIFICAR PAGAMENTO",
                })
    return pendentes


# ── exportação ────────────────────────────────────────────────────────────────

CAMPOS_SAIDA = ["MATRICULA", "VINCULO", "NOME", "OB", "MES_ANO", "VALOR"]
CAMPOS_PENDENTES = ["MATRICULA", "VINCULO", "NOME", "OB_ESTORNO", "MES_ANO", "VALOR_ESTORNADO", "SITUACAO"]


def gerar_csv_saida(linhas: list) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=CAMPOS_SAIDA)
    w.writeheader()
    w.writerows(linhas)
    return buf.getvalue().encode("utf-8-sig")


def gerar_csv_pendentes(linhas: list) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=CAMPOS_PENDENTES)
    w.writeheader()
    w.writerows(linhas)
    return buf.getvalue().encode("utf-8-sig")


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO OF — cruzamento OB → matrícula/vínculo (substitui a folha)
# ══════════════════════════════════════════════════════════════════════════════
#
# A base OF (export "PRDs_OBs") amarra cada OB de pagamento à matrícula e vínculo
# corretos (o vínculo efetivamente pago), além do PRD (num_despesa) e do mês.
# Colunas esperadas:
#   num_despesa | matricula | vinculo | cpf | num_ordem_bancaria | dt_pagamento | mes_pagamento
#
# Formato de OB difere do relatório de diárias:
#   relatório:  2026160101OB09096   (ano + UG + OB + seq)
#   OF:         2026OB09096          (ano + OB + seq, sem a UG)
# A normalização remove a UG para o cruzamento.

_OF_COLS = {
    "num_despesa":        ["num_despesa", "prd", "despesa"],
    "matricula":          ["matricula", "matrícula"],
    "vinculo":            ["vinculo", "vínculo"],
    "cpf":                ["cpf"],
    "num_ordem_bancaria": ["num_ordem_bancaria", "ordem_bancaria", "ob"],
    "dt_pagamento":       ["dt_pagamento", "data_pagamento", "data"],
    "mes_pagamento":      ["mes_pagamento", "mês_pagamento", "mes", "competencia", "competência"],
}


def normalizar_ob(ob):
    """Remove a UG entre o ano e 'OB' p/ casar os dois formatos. Caixa alta."""
    s = str(ob).strip().upper()
    m = re.match(r"^(\d{4})\d+(OB\w+)$", s)
    return (m.group(1) + m.group(2)) if m else s


def _mes_key(mes):
    """'5/2026' -> (2026, 5) para ordenação."""
    try:
        partes = re.split(r"[/\-]", str(mes).strip())
        if len(partes) >= 2:
            return (int(partes[1]), int(partes[0]))
    except (ValueError, IndexError):
        pass
    return (9999, 99)


def mes_display(mes):
    """'5/2026' -> '05/2026'."""
    try:
        partes = re.split(r"[/\-]", str(mes).strip())
        return f"{int(partes[0]):02d}/{int(partes[1])}"
    except (ValueError, IndexError):
        return str(mes).strip()


def _mapear_colunas_of(header):
    """Mapeia nome canônico -> índice na planilha, por correspondência de título."""
    idx = {}
    norm = [str(h).strip().lower().replace(" ", "_") for h in header]
    for canon, aliases in _OF_COLS.items():
        for a in aliases:
            if a in norm:
                idx[canon] = norm.index(a)
                break
    return idx


def ler_of(file_obj, nome_arquivo="of.xlsx"):
    """Lê a base OF (XLSX ou CSV) e devolve lista de dicts normalizados."""
    ext = nome_arquivo.rsplit(".", 1)[-1].lower()
    rows = ler_xlsx(file_obj) if ext in ("xlsx", "xls") else ler_csv_diaria(file_obj)
    if not rows:
        return []

    cols = _mapear_colunas_of(rows[0])
    if "num_ordem_bancaria" not in cols:
        return []  # sem coluna de OB não dá pra cruzar

    def g(r, c):
        i = cols.get(c)
        return str(r[i]).strip() if (i is not None and i < len(r)) else ""

    registros = []
    for r in rows[1:]:
        ob = g(r, "num_ordem_bancaria")
        if not ob:
            continue
        registros.append({
            "prd":       g(r, "num_despesa"),
            "matricula": g(r, "matricula"),
            "vinculo":   g(r, "vinculo"),
            "cpf":       g(r, "cpf").zfill(11) if g(r, "cpf").isdigit() else g(r, "cpf"),
            "ob":        ob,
            "ob_norm":   normalizar_ob(ob),
            "dt":        g(r, "dt_pagamento"),
            "mes":       g(r, "mes_pagamento"),
        })
    return registros


def indexar_of(of_rows):
    """dict ob_norm -> lista de registros OF (um OB pode, raramente, repetir)."""
    idx = {}
    for r in of_rows:
        idx.setdefault(r["ob_norm"], []).append(r)
    return idx


def meses_da_of(of_rows):
    """Lista de meses presentes na OF, ordenada, no formato 'MM/YYYY'."""
    vistos = {}
    for r in of_rows:
        if r["mes"]:
            vistos[mes_display(r["mes"])] = _mes_key(r["mes"])
    return [m for m, _ in sorted(vistos.items(), key=lambda kv: kv[1])]


def of_para_csv(of_rows):
    """Serializa a OF para CSV (persistência em disco)."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["num_despesa", "matricula", "vinculo", "cpf",
                "num_ordem_bancaria", "dt_pagamento", "mes_pagamento"])
    for r in of_rows:
        w.writerow([r["prd"], r["matricula"], r["vinculo"], r["cpf"],
                    r["ob"], r["dt"], r["mes"]])
    return buf.getvalue().encode("utf-8-sig")


# ── cruzamento híbrido (OB exato → fallback CPF) ──────────────────────────────

CAMPOS_OF_SAIDA = ["MATRICULA", "VINCULO", "NOME", "CPF", "OB", "MÊS/ANO", "VALOR", "PRD", "ORIGEM"]
CAMPOS_OF_CADASTRO = ["OB", "NOME", "CPF", "DATA", "VALOR",
                      "MATRICULA_PROVAVEL", "VINCULO_PROVAVEL"]
CAMPOS_OF_NAO_ENCONTRADAS = ["OB", "NOME", "CPF", "DATA", "VALOR", "MOTIVO"]


def _norm_cpf(cpf):
    c = str(cpf).strip()
    return c.zfill(11) if (c.isdigit() and len(c) <= 11) else c


def indexar_of_por_cpf(of_rows):
    """
    dict cpf -> (matricula, vinculo). A OF só registra o vínculo efetivamente
    pago, então cada CPF mapeia para um único vínculo (validado: 0 ambíguos).
    """
    idx = {}
    for r in of_rows:
        idx.setdefault(r["cpf"], (r["matricula"], r["vinculo"]))
    return idx


def cruzar_of_hibrido(registros, of_rows, mes_alvo):
    """
    Cruzamento híbrido:
      1. casa por OB (exato, por pagamento) — usando toda a OF;
      2. se a OB não estiver na OF, infere matrícula/vínculo pelo CPF;
      3. se o CPF também não estiver na OF, fica sem vínculo.

    `mes_alvo` ('MM/YYYY') é apenas o rótulo da competência na saída.

    Devolve (vinculados, a_cadastrar, nao_encontradas):
      • vinculados     — linhas resolvidas (ORIGEM = "OB" ou "CPF (inferido)").
      • a_cadastrar    — OBs ausentes da OF MAS com vínculo recuperado por CPF
                         (o time de cadastro deve registrá-las na OF).
      • nao_encontradas — OBs sem matrícula/vínculo: nem a OB nem o CPF constam
                         na OF (precisam de investigação).
    """
    ob_idx = indexar_of(of_rows)
    cpf_idx = indexar_of_por_cpf(of_rows)

    vinculados = {}        # dedup por ob_norm
    a_cadastrar = {}       # B: ausente da OF, vínculo por CPF
    nao_encontradas = {}   # C: nem OB nem CPF na OF
    for reg in registros:
        ob_raw = reg["OB"]
        ob_norm = normalizar_ob(ob_raw)
        cpf = _norm_cpf(reg.get("CPF_CNPJ", ""))
        nome = reg.get("NOME", "")
        valor = reg.get("VALOR", "")
        data = reg.get("DATA", "")

        infos = ob_idx.get(ob_norm)
        if infos:
            # OB consta na OF: prefere registro do mês alvo, senão o primeiro
            info = next((x for x in infos if mes_display(x["mes"]) == mes_alvo), infos[0])
            if ob_norm not in vinculados:
                vinculados[ob_norm] = {
                    "MATRICULA": info["matricula"], "VINCULO": info["vinculo"],
                    "NOME": nome, "CPF": cpf, "OB": ob_raw, "MÊS/ANO": mes_alvo,
                    "VALOR": valor, "PRD": info["prd"], "ORIGEM": "OB",
                }
        else:
            mv = cpf_idx.get(cpf)
            if mv:
                # B — vínculo recuperado por CPF; OB ainda precisa entrar na OF
                if ob_norm not in vinculados:
                    vinculados[ob_norm] = {
                        "MATRICULA": mv[0], "VINCULO": mv[1], "NOME": nome,
                        "CPF": cpf, "OB": ob_raw, "MÊS/ANO": mes_alvo,
                        "VALOR": valor, "PRD": "", "ORIGEM": "CPF (inferido)",
                    }
                if ob_norm not in a_cadastrar:
                    a_cadastrar[ob_norm] = {
                        "OB": ob_raw, "NOME": nome, "CPF": cpf, "DATA": data,
                        "VALOR": valor, "MATRICULA_PROVAVEL": mv[0],
                        "VINCULO_PROVAVEL": mv[1],
                    }
            else:
                # C — não encontrada: nem OB nem CPF na OF (sem matrícula/vínculo)
                if ob_norm not in nao_encontradas:
                    nao_encontradas[ob_norm] = {
                        "OB": ob_raw, "NOME": nome, "CPF": cpf, "DATA": data,
                        "VALOR": valor,
                        "MOTIVO": "OB e CPF não localizados na OF",
                    }

    enc = sorted(vinculados.values(), key=lambda x: x["NOME"])
    cad = sorted(a_cadastrar.values(), key=lambda x: x["NOME"])
    nao = sorted(nao_encontradas.values(), key=lambda x: x["NOME"])
    return enc, cad, nao


def gerar_csv_of_saida(linhas):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=CAMPOS_OF_SAIDA)
    w.writeheader()
    w.writerows(linhas)
    return buf.getvalue().encode("utf-8-sig")


def gerar_csv_of_cadastro(linhas):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=CAMPOS_OF_CADASTRO)
    w.writeheader()
    w.writerows(linhas)
    return buf.getvalue().encode("utf-8-sig")


def gerar_csv_of_nao_encontradas(linhas):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=CAMPOS_OF_NAO_ENCONTRADAS)
    w.writeheader()
    w.writerows(linhas)
    return buf.getvalue().encode("utf-8-sig")


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO RESSARCIMENTO
# ══════════════════════════════════════════════════════════════════════════════
#
# Layout do relatório SIAFE de ressarcimento (índices 0-based):
#   0: NOME DO CREDOR (com CPF: "FULANO - 12345678901")
#   1: (vazia)
#   2: SITUAÇÃO        ("ENVIADO", "ANULADO", ... / vazia em linhas de subtotal)
#   3: DOCUMENTO (OB)  ("2026160101OB10440")
#   4: NÚMERO DO PROCESSO
#   5: DATA
#   6: EVENTO
#   7: FONTE DE RECURSO
#   8: VALOR

import unicodedata

# palavra-chave que identifica uma diária na descrição.
# o match é insensível a acento e maiúsculas; "DIARIA" também cobre "DIARIAS".
PALAVRA_DIARIA = "DIARIA"
SITUACAO_ANULADO_RESS = "ANULADO"


def _sem_acento(texto):
    """Remove acentos e devolve em maiúsculas para comparação robusta."""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    sem = "".join(c for c in nfkd if not unicodedata.combining(c))
    return sem.upper()

_HEADER_PADRAO_RESS = [
    "NOME DO CREDOR", "", "SITUAÇÃO", "DOCUMENTO", "NÚMERO DO PROCESSO",
    "DATA", "EVENTO", "FONTE DE RECURSO", "VALOR", "DESCRIÇÃO",
]


def _valor_numerico(v):
    """Tenta converter para float; devolve None se não der."""
    s = str(v).strip().replace("\n", "")
    if not s:
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def ler_ressarcimento_xlsx(file_obj, nome_arquivo="arquivo.xlsx"):
    """
    Lê o relatório de ressarcimento (XLSX ou CSV) e extrai as OBs válidas.
    Ignora cabeçalho e linhas de subtotal (SITUAÇÃO vazia).
    Devolve (registros, header) onde cada registro é um dict.
    """
    ext = nome_arquivo.rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xls"):
        rows = ler_xlsx(file_obj)
    else:
        rows = ler_csv_diaria(file_obj)

    header = _padrow(rows[0], 9) if rows else []
    registros = []
    for r in rows:
        r = _padrow(r, 9)
        situacao = r[2].strip()
        ob = r[3].strip()

        # ignora cabeçalho, subtotais e linhas sem OB
        if not situacao or situacao.upper() in SITUACOES_IGNORAR:
            continue
        if r[0].strip().upper() == "NOME DO CREDOR":
            continue
        if not ob:
            continue
        if _valor_numerico(r[8]) is None:
            continue

        registros.append({
            "nome":     r[0].strip(),
            "situacao": situacao,
            "ob":       ob,
            "processo": r[4].strip(),
            "data":     r[5].strip(),
            "evento":   r[6].strip(),
            "fonte":    r[7].strip(),
            "valor":    r[8].strip(),
            "_row":     r[:9],   # linha original p/ regerar o relatório
        })
    return registros, header


def obs_unicas(registros):
    """Lista de OBs distintas, preservando ordem de aparição."""
    vistas = []
    seen = set()
    for reg in registros:
        ob = reg["ob"]
        if ob not in seen:
            seen.add(ob)
            vistas.append(ob)
    return vistas


def classificar_descricao(descricao):
    """
    Devolve o status de revisão de uma descrição capturada:
      'ok'        -> contém 'DIÁRIA' / variações (auto-confirmada)
      'verificar' -> tem texto, mas sem 'DIÁRIA' (precisa de checkbox)
      'manual'    -> sem descrição (precisa digitar manualmente)
    """
    d = (descricao or "").strip()
    if not d:
        return "manual"
    if PALAVRA_DIARIA in _sem_acento(d):
        return "ok"
    return "verificar"


def filtrar_ressarcimento(registros, decisoes):
    """
    Aplica o filtro final.
    `decisoes`: dict ob -> {"descricao": str, "is_diaria": bool}

    O relatório é, por natureza, de ressarcimento — então o critério é se o
    lançamento é, de fato, uma DIÁRIA (descrição contém "diária"/variações).

    Remove um registro se:
      • SITUAÇÃO == ANULADO, OU
      • valor negativo (estorno), OU
      • não for marcado como diária.

    Devolve (mantidos, removidos), com a descrição anexada a cada item.
    """
    mantidos, removidos = [], []
    for reg in registros:
        dec = decisoes.get(reg["ob"], {})
        descricao = dec.get("descricao", "")
        is_diaria = bool(dec.get("is_diaria", False))
        anulado = reg["situacao"].strip().upper() == SITUACAO_ANULADO_RESS
        valor = _valor_numerico(reg["valor"])
        negativo = valor is not None and valor < 0

        reg_out = dict(reg)
        reg_out["descricao"] = descricao

        if anulado:
            motivo = "ANULADO"
        elif negativo:
            motivo = "VALOR NEGATIVO"
        elif not is_diaria:
            motivo = "NÃO É DIÁRIA"
        else:
            motivo = None

        if motivo:
            reg_out["motivo_remocao"] = motivo
            removidos.append(reg_out)
        else:
            mantidos.append(reg_out)
    return mantidos, removidos


def gerar_xlsx_ressarcimento(registros, header=None):
    """
    Gera o relatório de ressarcimento filtrado em XLSX, preservando o layout
    original (9 colunas) + DESCRIÇÃO ao final, para reuso na aba Processamento.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RESSARCIMENTO"

    cab = list(header) if header else _HEADER_PADRAO_RESS[:9]
    cab = _padrow(cab, 9)[:9] + ["DESCRIÇÃO"]
    ws.append(cab)

    for reg in registros:
        linha = list(reg.get("_row", []))[:9]
        linha = _padrow(linha, 9)[:9]
        linha.append(reg.get("descricao", ""))
        ws.append(linha)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()
